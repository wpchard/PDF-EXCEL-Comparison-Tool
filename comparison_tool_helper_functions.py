#helper functions for the TFFF comparison tool. these functions perform specific tasks such as converting PDFs to Excel, normalizing serial numbers, fixing common formatting issues from PDF imports, and determining the status of each entry based on merge results and date comparisons. they are designed to be called by the main functions in comparison_tool_functions.py to keep the code organized and modular.
import pandas as pd
import datetime
from tkinter.filedialog import askopenfilename
import camelot
import os
from tkinter import Tk
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
#define some commonly used variables that will be used in multiple functions
EXPECTED_COLUMNS = ['DATE RECEIVED', 'SERIAL NO.', 'CCAM', 'CUST #', 'CUSTOMER NAME', 'ORIG. WO#', 'LAST SCAN DATE', 'LAST SCAN DESCRIPTION']
merge_key = "SERIAL NO."
date_column = "LAST SCAN DATE"
old_date_col = f"{date_column}_OLD"


# This function converts a PDF file to Excel format by extracting tables using Camelot. It prompts the user to select a PDF file, processes the tables to extract relevant information based on fixed character positions, and saves the cleaned data to a new Excel file in the same directory as the original PDF. The function returns the path to the newly created Excel file for further processing.
def convert_pdf_to_excel(tfff):
    Tk().withdraw()
    pdf_path = tfff
    tables = camelot.read_pdf(
        pdf_path,
        pages="all",
        flavor="stream"
    )
    rows = []
    for table in tables:
        df = table.df
        if df.empty:
            continue
        for row in table.df[0]:
            # Skip blank rows
            if not row.strip():
                continue
            # Skip short rows
            if len(row) < 100:
                continue
            parsed = {
                "DATE RECEIVED": row[0:12].strip(),
                "SERIAL NO.": row[13:23].strip(),
                "CCAM": row[26:31].strip(),
                "CUST #": row[36:41].strip(),
                "CUSTOMER NAME": row[44:67].strip(),
                "ORIG. WO#": row[71:77].strip(),
                "LAST SCAN DATE": row[80:93].strip(),
                "LAST SCAN DESCRIPTION": row[97:125].strip(),
            }
            rows.append(parsed)
    final_df = pd.DataFrame(rows)
    
    #generate file name
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    excel_path = os.path.join(
        os.path.dirname(pdf_path),
        f"{base_name}.xlsx"
    )
    #save to excel and return path
    final_df.to_excel(excel_path, index=False)
    
    # Auto-fit column widths
    _auto_fit_columns_helper(excel_path)
    
    print(f"\nCreated: {excel_path}")
    return excel_path

def _auto_fit_columns_helper(file_path):
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set column width with some padding (1.2 multiplier for better readability)
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(file_path)

#gets excel file path from the user
def get_file_location(label="input"):
    Tk().withdraw()
    print(f"Select {label} PDF or Excel file:")
    file_path = askopenfilename(
        filetypes=[("All Supported", "*.pdf *.xlsx"), ("PDF Files", "*.pdf"), ("Excel Files", "*.xlsx")]
    )
    if not file_path:
        print("No file selected.")
        exit()
    return file_path

# This function checks if a given value is a valid date. It handles various formats and ensures that the date falls within a reasonable range (2020-2035). It also filters out values that are likely to be non-date text, such as those starting with letters or empty strings.
def is_valid_date(value):

    if pd.isna(value):
        return False

    if isinstance(value, (pd.Timestamp, datetime.datetime)):
        return True

    value = str(value).strip()

    if not value or value[0].isalpha():
        return False

    parsed = pd.to_datetime(value, errors='coerce')

    return (
        pd.notna(parsed)
        and 2020 <= parsed.year <= 2035
    )

# This function checks if the DataFrame has only one column, which is a common issue when importing from PDFs. If it detects a single-column format, it attempts to split the column into multiple columns based on common delimiters (multiple spaces or tabs). If the split results in more than one column, it assigns new column indices and returns the modified DataFrame. If not, it returns the original DataFrame unchanged.
def fix_single_column(df, sheet_name):
    #check if only single column exists, if so split. other issues with different columns are fixed in other functions.
    if len(df.columns) != 1:
        return df
    split_df = df[0].astype(str).str.split(
        r'\s{2,}|\t',
        expand=True
    )
    if len(split_df.columns) > 1:
        split_df.columns = range(len(split_df.columns))
        return split_df
    return df

# This function checks if the first column of the DataFrame appears to be a row number column (numeric values starting from 1 and less than 500). If it detects such a column, it drops it and reassigns the column indices. This is common in TFFF reports where the first column may contain row numbers that are not needed for analysis.
def remove_row_number_column(df, sheet_name):
    if len(df) == 0 or len(df.columns) == 0:
        return df
    try:
        first_col = pd.to_numeric(
            df[0].dropna(),
            errors='coerce'
        )
        if (
            first_col.notna().all()
            and first_col.max() < 500
        ):
            print(
                f"  Sheet '{sheet_name}': "
                f"Dropping row-number column"
            )
            df = df.drop(columns=[0])
            df.columns = range(len(df.columns))
    except:
        pass
    return df

# This function is designed to clean and normalize a single serial number in the DataFrame. 
def normalize_serial(value):
    if pd.isna(value):
        return None
    
    # Convert to string first
    val_str = str(value).strip()
    
    # Remove common artifacts from PDF conversion
    val_str = val_str.replace('\n', '').replace('\r', '').replace('\t', '')
    val_str = val_str.replace('\u00A0', '')  # non-breaking space
    val_str = val_str.replace(',', '')        # thousand separators
    val_str = val_str.replace(' ', '')
    
    # Handle float strings like "12345.0" -> "12345"
    try:
        num = float(val_str)
        if num.is_integer():
            return str(int(num))
        else:
            return str(num)
    except ValueError:
        # Not a number, return cleaned string
        return val_str.upper()

#fixes blank spaces for new entries in scan description, or descriptions with only partial information
def normalize_scan_description(df):

    if 'LAST SCAN DESCRIPTION' not in df.columns:
        return df

    # Convert to string safely
    scan_col = (df['LAST SCAN DESCRIPTION'].fillna('').astype(str).str.strip())

    # Detect blanks or NO SCANS text
    no_scan_mask = ((scan_col == '')|(scan_col.str.contains(r'NO\s*SCANS',case=False,na=False)))

    # Standardize value
    df.loc[no_scan_mask, 'LAST SCAN DESCRIPTION'] = 'NO SCANS FOUND FOR THIS UNIT'

    return df

#checks the merge status and compares dates to determine if it's a new entry, removed, unchanged, or new scan. This is applied to each row of the merged dataframe.
def determine_status(row):
    if row['_merge'] == 'left_only':
        return 'New Entry'
    elif row['_merge'] == 'right_only':
        return 'Removed'
    elif row['_merge'] == 'both':
        old_date = row[old_date_col]
        new_date = row[date_column]
        
        # Both dates are NaT = no change
        if pd.isna(old_date) and pd.isna(new_date):
            return 'Unchanged'
        
        # Old date missing but new date exists = new scan
        if pd.isna(old_date) and pd.notna(new_date):
            return 'New Scan'
        
        # New date missing but old date existed = unlikely but unchanged
        if pd.notna(old_date) and pd.isna(new_date):
            return 'Unchanged'
        
        # Both dates exist - compare them
        if new_date > old_date:
            return 'New Scan'      
        return 'Unchanged'
    
    #catch strange cases
    return 'Unknown'

def auto_fit_columns(file_path, sheet_name):
    from openpyxl import load_workbook
    
    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name]
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set column width with some padding (1.2 multiplier for better readability)
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(file_path)